"""
FastAPI Backend - Chat endpoint with Claude API + MCP tool_use loop.

Proxies between the React frontend and Claude API, forwarding tool calls
to the BSL MCP server. Includes chat logging, usage tracking, rate limiting,
error handling, and content serving.
"""

import json
import logging
import os
import time
import uuid
from datetime import date
from pathlib import Path

import structlog
from dotenv import load_dotenv

load_dotenv(Path(__file__).parent / ".env")

import anthropic  # noqa: E402
from fastapi import Depends, FastAPI, HTTPException, Request  # noqa: E402
from fastapi.middleware.cors import CORSMiddleware  # noqa: E402
from fastapi.staticfiles import StaticFiles  # noqa: E402
from pydantic import BaseModel  # noqa: E402
from starlette.middleware.base import BaseHTTPMiddleware  # noqa: E402

from . import config as app_config  # noqa: E402
from .auth import require_auth  # noqa: E402
from .permissions import get_user_permissions, can_access_dashboard, get_dashboard_registry, is_admin  # noqa: E402
from .chat_logger import ChatLogger  # noqa: E402
from .dashboard import load_dashboard_config, execute_dashboard, validate_all_dashboards, validate_dashboard_config  # noqa: E402
from .event_logger import create_logger as create_event_logger  # noqa: E402
from .dashboard_models import DashboardQueryRequest, DashboardQueryResponse, PageRequest, PageResponseItem  # noqa: E402
from .rate_limiter import RateLimiter  # noqa: E402
from .usage_logger import UsageLoggingMiddleware  # noqa: E402

try:
    from .semantic import MCP_TOOLS, call_tool as semantic_call_tool  # noqa: E402
    _SEMANTIC_AVAILABLE = True
except ImportError as _sem_err:
    logging.getLogger(__name__).warning(f"Semantic layer unavailable: {_sem_err}")
    MCP_TOOLS = []

    async def semantic_call_tool(name, input):
        return '{"error": "Semantic layer unavailable — boring_semantic_layer dependency broken"}'

    _SEMANTIC_AVAILABLE = False

logger = logging.getLogger(__name__)

app = FastAPI(title="Tanit")

# CORS — reads from ALLOWED_ORIGINS env var (comma-separated), falls back to localhost
_default_origins = ["http://localhost:3000", "http://localhost:3002", "http://localhost:5173"]
_env_origins = os.environ.get("ALLOWED_ORIGINS", "")
_origins = [o.strip() for o in _env_origins.split(",") if o.strip()] if _env_origins else _default_origins

app.add_middleware(
    CORSMiddleware,
    allow_origins=_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configuration
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
MODEL = "claude-haiku-4-5-20251001"

# Load system prompt + inject metric definitions + dashboard schema
SYSTEM_PROMPT_PATH = Path(__file__).parent / "prompts" / "system.md"
SYSTEM_PROMPT = SYSTEM_PROMPT_PATH.read_text()

_metric_defs_path = app_config.CONTENT_ROOT / "semantic-layer" / "metric-definitions.md"
if _metric_defs_path.exists():
    SYSTEM_PROMPT += "\n\n## Metric Definitions\n\n" + _metric_defs_path.read_text()

_dashboard_schema_path = Path(__file__).parent / "prompts" / "dashboard_schema.md"
if _dashboard_schema_path.exists():
    SYSTEM_PROMPT += "\n\n" + _dashboard_schema_path.read_text()

# In-memory session storage
sessions: dict[str, list] = {}

# Claude client
claude_client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

# Chat logger (in-memory only)
chat_logger = ChatLogger()

# Rate limiter
rate_limiter = RateLimiter()

# Event logger (GCS or local, configured via EVENT_LOG_BACKEND)
try:
    event_logger = create_event_logger()
except Exception as _ev_err:
    logger.warning(f"Event logger init failed, using no-op logger: {_ev_err}")

    class _NoopLogger:
        def log(self, event_type: str, payload: dict) -> None:
            pass

    event_logger = _NoopLogger()

# Usage middleware — emits page_view/user_session events via event_logger
app.add_middleware(UsageLoggingMiddleware, event_logger=event_logger)


class ChatRequest(BaseModel):
    message: str
    session_id: str | None = None


class ChatResponse(BaseModel):
    response: str
    session_id: str
    tool_calls: list[dict] | None = None
    chart_spec: dict | None = None


class SemanticQueryRequest(BaseModel):
    model_name: str
    dimensions: list[str] = []
    measures: list[str] = []
    filters: list[str] = []
    limit: int = 100


def build_chart_spec(tool_name: str, tool_input: dict, tool_result: str) -> dict | None:
    """
    Generate a chart spec based on query shape.

    Heuristic mapping:
    - 1 time dim + 1 measure → line chart
    - 1 time dim + 1 measure + 1 categorical dim → grouped line chart
    - 1 categorical dim + 1 measure → bar chart
    - 0 dims + 1 measure → big_value
    - 0 dims + 2+ measures → multiple big_values (first measure for now)
    - Anything else → None
    """
    if tool_name != "query_model":
        return None

    try:
        # Parse result to get data
        result_data = json.loads(tool_result)
        if "error" in result_data or not result_data.get("data"):
            return None

        dimensions = tool_input.get("dimensions", [])
        measures = tool_input.get("measures", [])
        data = result_data["data"]

        if not data or not measures:
            return None

        # Identify time dimensions (common patterns)
        time_dimensions = [
            d
            for d in dimensions
            if any(
                time_word in d.lower()
                for time_word in [
                    "date",
                    "time",
                    "month",
                    "year",
                    "day",
                    "week",
                    "quarter",
                ]
            )
        ]
        categorical_dimensions = [d for d in dimensions if d not in time_dimensions]

        # 0 dims + 1 measure → big_value
        if len(dimensions) == 0 and len(measures) >= 1:
            measure = measures[0]
            value = data[0].get(measure) if data else 0
            return {
                "type": "big_value",
                "value": value,
                "label": measure.replace("_", " ").title(),
                "format": "auto",
            }

        # 1 time dim + 1 measure → line chart
        if (
            len(time_dimensions) == 1
            and len(measures) == 1
            and len(categorical_dimensions) == 0
        ):
            time_dim = time_dimensions[0]
            measure = measures[0]
            return {
                "type": "line",
                "title": f"{measure.replace('_', ' ').title()} Over Time",
                "x": time_dim,
                "y": measure,
                "data": data,
            }

        # 1 time dim + 1 measure + 1 categorical dim → grouped line chart
        if (
            len(time_dimensions) == 1
            and len(measures) == 1
            and len(categorical_dimensions) == 1
        ):
            time_dim = time_dimensions[0]
            measure = measures[0]
            category = categorical_dimensions[0]
            return {
                "type": "line",
                "title": f"{measure.replace('_', ' ').title()} by {category.replace('_', ' ').title()}",
                "x": time_dim,
                "y": measure,
                "series": category,
                "data": data,
            }

        # 1 categorical dim + 1 measure → bar chart
        if (
            len(dimensions) == 1
            and len(measures) == 1
            and len(categorical_dimensions) == 1
        ):
            category = categorical_dimensions[0]
            measure = measures[0]
            return {
                "type": "bar",
                "title": f"{measure.replace('_', ' ').title()} by {category.replace('_', ' ').title()}",
                "x": category,
                "y": measure,
                "data": data,
            }

        # Default: no chart for complex queries
        return None

    except (json.JSONDecodeError, KeyError, IndexError):
        return None


@app.post("/api/chat")
async def chat(
    request: ChatRequest, claims: dict = Depends(require_auth)
) -> ChatResponse:
    """Handle a chat message with tool_use loop."""
    if not ANTHROPIC_API_KEY:
        raise HTTPException(status_code=500, detail="ANTHROPIC_API_KEY not set")

    user_id = claims.get("sub", "unknown")
    user_email = claims.get("email", "")

    # Rate limiting
    allowed, reason = rate_limiter.check(user_id)
    if not allowed:
        raise HTTPException(status_code=429, detail=reason)

    # Get or create session (restore from SQLite if needed)
    session_id = request.session_id or str(uuid.uuid4())
    if session_id not in sessions:
        restored = chat_logger.load_session(session_id)
        sessions[session_id] = restored if restored else []

    # Log session
    chat_logger.upsert_session(session_id, user_id, user_email)

    messages = sessions[session_id]

    # Add user message
    messages.append({"role": "user", "content": request.message})

    # Log user message
    chat_logger.log_message(session_id, "user", request.message)

    # Build system prompt with current date
    system = f"{SYSTEM_PROMPT}\n\nToday's date: {date.today().isoformat()}"

    # Tool use loop
    tool_calls_log = []
    chart_spec = None
    max_iterations = 10
    total_tokens_in = 0
    total_tokens_out = 0

    try:
        for _ in range(max_iterations):
            try:
                response = claude_client.messages.create(
                    model=MODEL,
                    max_tokens=4096,
                    system=system,
                    messages=messages,
                    tools=MCP_TOOLS,
                )
            except anthropic.AuthenticationError:
                logger.error("Anthropic API authentication failed")
                return ChatResponse(
                    response="Chat service configuration error. Contact admin.",
                    session_id=session_id,
                )
            except anthropic.RateLimitError:
                return ChatResponse(
                    response="Chat service is busy. Please wait a moment and try again.",
                    session_id=session_id,
                )
            except anthropic.APIError as e:
                logger.error(f"Anthropic API error: {e}")
                return ChatResponse(
                    response="An error occurred with the chat service. Please try again.",
                    session_id=session_id,
                )

            # Track token usage
            if hasattr(response, "usage"):
                total_tokens_in += response.usage.input_tokens
                total_tokens_out += response.usage.output_tokens

            # Check if response contains tool use
            has_tool_use = any(block.type == "tool_use" for block in response.content)

            if not has_tool_use or response.stop_reason == "end_turn":
                # Extract text response
                text_parts = []
                for block in response.content:
                    if block.type == "text":
                        text_parts.append(block.text)

                final_text = "\n".join(text_parts)
                messages.append({"role": "assistant", "content": response.content})

                # Log assistant response
                chat_logger.log_message(
                    session_id,
                    "assistant",
                    final_text,
                    tool_calls=tool_calls_log if tool_calls_log else None,
                    chart_spec=chart_spec,
                    tokens_in=total_tokens_in,
                    tokens_out=total_tokens_out,
                )

                # Record rate limit
                rate_limiter.record(user_id)

                return ChatResponse(
                    response=final_text,
                    session_id=session_id,
                    tool_calls=tool_calls_log if tool_calls_log else None,
                    chart_spec=chart_spec,
                )

            # Process tool calls
            assistant_content = []
            tool_results = []

            for block in response.content:
                if block.type == "text":
                    assistant_content.append({"type": "text", "text": block.text})
                elif block.type == "tool_use":
                    assistant_content.append(
                        {
                            "type": "tool_use",
                            "id": block.id,
                            "name": block.name,
                            "input": block.input,
                        }
                    )

                    # Execute tool via embedded semantic layer
                    try:
                        result = await semantic_call_tool(block.name, block.input)
                    except Exception as e:
                        logger.error(f"MCP tool call failed: {e}")
                        result = json.dumps(
                            {
                                "error": "Data service is temporarily unavailable. Try again in a few minutes."
                            }
                        )

                    tool_calls_log.append(
                        {
                            "tool": block.name,
                            "input": block.input,
                            "result_preview": result[:200]
                            if len(result) > 200
                            else result,
                        }
                    )

                    # Generate chart spec for query_model calls
                    spec = build_chart_spec(block.name, block.input, result)
                    if spec:
                        chart_spec = spec

                    tool_results.append(
                        {
                            "type": "tool_result",
                            "tool_use_id": block.id,
                            "content": result,
                        }
                    )

            # Add assistant message and tool results
            messages.append({"role": "assistant", "content": assistant_content})
            messages.append({"role": "user", "content": tool_results})

    except Exception as e:
        logger.error(f"Unexpected error in chat: {e}")
        return ChatResponse(
            response="An unexpected error occurred. Please try again.",
            session_id=session_id,
            tool_calls=tool_calls_log if tool_calls_log else None,
        )

    # If we hit max iterations
    final_text = "I wasn't able to complete the analysis within the allowed number of steps. Please try a more specific question."
    chat_logger.log_message(
        session_id,
        "assistant",
        final_text,
        tool_calls=tool_calls_log,
        chart_spec=chart_spec,
        tokens_in=total_tokens_in,
        tokens_out=total_tokens_out,
    )
    rate_limiter.record(user_id)

    return ChatResponse(
        response=final_text,
        session_id=session_id,
        tool_calls=tool_calls_log,
        chart_spec=chart_spec,
    )


@app.post("/api/semantic/query")
async def semantic_query(
    request: SemanticQueryRequest, claims: dict = Depends(require_auth)
):
    """Forward a semantic query to the BSL MCP server and return rows."""
    user_id = claims.get("sub", "unknown")

    # Rate limiting
    allowed, reason = rate_limiter.check(user_id)
    if not allowed:
        raise HTTPException(status_code=429, detail=reason)

    tool_input = {
        "model_name": request.model_name,
        "dimensions": request.dimensions,
        "measures": request.measures,
        "filters": request.filters,
        "limit": request.limit,
    }
    result_str = await semantic_call_tool("query_model", tool_input)
    try:
        result = json.loads(result_str)
    except json.JSONDecodeError:
        raise HTTPException(status_code=502, detail="Invalid response from MCP server")
    if "error" in result:
        raise HTTPException(status_code=502, detail=result["error"])
    rate_limiter.record(user_id)
    return {"rows": result.get("rows", result.get("data", []))}


@app.get("/api/health")
async def health():
    """Deep health check with dependency status."""
    checks = {
        "anthropic_api": "ok",
        "content_files": "ok",
        "duckdb": "ok",
    }

    # Check Anthropic API key (basic validation)
    if not ANTHROPIC_API_KEY:
        checks["anthropic_api"] = "not_configured"

    # Check content files
    _content_data_dir = app_config.CONTENT_ROOT / "data"
    if (
        not _content_data_dir.exists()
        or not (_content_data_dir / "mockMetrics.json").exists()
    ):
        checks["content_files"] = "missing"

    # Check DuckDB engine
    try:
        from .duckdb_engine import get_duckdb_engine
        engine = get_duckdb_engine()
        if not engine._tables:
            checks["duckdb"] = "no_tables"
        else:
            checks["duckdb"] = "ok"
    except Exception:
        checks["duckdb"] = "not_initialized"

    status = "ok" if all(v == "ok" for v in checks.values()) else "degraded"

    return {"status": status, **checks}


# ---------------------------------------------------------------------------
# Dashboard engine endpoints (config-driven query backend)
# ---------------------------------------------------------------------------


@app.post("/api/dashboard/{slug}/query")
async def dashboard_query(
    slug: str,
    request: DashboardQueryRequest,
    claims: dict = Depends(require_auth),
) -> DashboardQueryResponse:
    """Execute dashboard queries and return results."""
    user_id = claims.get("sub", "unknown")

    # Rate limiting
    allowed, reason = rate_limiter.check(user_id)
    if not allowed:
        raise HTTPException(status_code=429, detail=reason)

    # Permission check
    perms = get_user_permissions(claims)
    if not can_access_dashboard(perms, slug):
        raise HTTPException(status_code=403, detail="Dashboard access denied")

    # Load config
    try:
        config = load_dashboard_config(slug)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"Dashboard '{slug}' not found")

    # Find the dashboard directory for SQL file resolution
    dashboard_dir = app_config.dashboards_dir() / slug

    # Execute
    result = await execute_dashboard(
        config=config,
        filter_values=request.filters,
        pagination=request.pagination,
        data_source_filter=request.data_source,
        dashboard_dir=dashboard_dir,
    )

    rate_limiter.record(user_id)
    return result


@app.post("/api/dashboard/{slug}/export/{source_id}/csv")
async def dashboard_export_csv(
    slug: str,
    source_id: str,
    request: DashboardQueryRequest,
    claims: dict = Depends(require_auth),
):
    """Export a data source as CSV with streaming response."""
    import csv
    import io
    from starlette.responses import StreamingResponse

    user_id = claims.get("sub", "unknown")
    allowed, reason = rate_limiter.check(user_id)
    if not allowed:
        raise HTTPException(status_code=429, detail=reason)

    perms = get_user_permissions(claims)
    if not can_access_dashboard(perms, slug):
        raise HTTPException(status_code=403, detail="Dashboard access denied")

    try:
        config = load_dashboard_config(slug)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"Dashboard '{slug}' not found")

    if source_id not in config.data_sources:
        raise HTTPException(status_code=400, detail=f"Unknown data source: {source_id}")

    dashboard_dir = app_config.dashboards_dir() / slug

    result = await execute_dashboard(
        config=config,
        filter_values=request.filters,
        pagination={},
        data_source_filter=source_id,
        dashboard_dir=dashboard_dir,
    )

    ds = result.data_sources.get(source_id, {})
    if isinstance(ds, dict) and ds.get("status") == "error":
        raise HTTPException(status_code=502, detail=ds.get("error", "Query failed"))

    rows = ds.get("rows", []) if isinstance(ds, dict) else []
    if len(rows) > 100_000:
        raise HTTPException(status_code=400, detail="Export limited to 100,000 rows. Use filters to narrow results.")

    # Get column display names from config layout
    col_defs = {}
    from .dashboard import _collect_panel_columns
    for node in config.layout:
        if node.data_source == source_id and node.columns:
            for col in node.columns:
                col_defs[col.key] = col
            break

    columns = ds.get("columns", []) if isinstance(ds, dict) else []

    def generate():
        output = io.StringIO()
        writer = csv.writer(output)
        # Header row with display names
        header = [col_defs[c].label if c in col_defs else c for c in columns]
        writer.writerow(header)
        yield output.getvalue()
        output.seek(0)
        output.truncate(0)
        # Data rows
        for row in rows:
            writer.writerow([row.get(c, "") for c in columns])
            yield output.getvalue()
            output.seek(0)
            output.truncate(0)

    rate_limiter.record(user_id)
    filename = f"{slug}_{source_id}.csv"
    return StreamingResponse(
        generate(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/dashboard/{slug}/export/{source_id}/xlsx")
async def dashboard_export_xlsx(
    slug: str,
    source_id: str,
    request: DashboardQueryRequest,
    claims: dict = Depends(require_auth),
):
    """Export a data source as Excel XLSX."""
    import io
    from starlette.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font

    user_id = claims.get("sub", "unknown")
    allowed, reason = rate_limiter.check(user_id)
    if not allowed:
        raise HTTPException(status_code=429, detail=reason)

    perms = get_user_permissions(claims)
    if not can_access_dashboard(perms, slug):
        raise HTTPException(status_code=403, detail="Dashboard access denied")

    try:
        config = load_dashboard_config(slug)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"Dashboard '{slug}' not found")

    if source_id not in config.data_sources:
        raise HTTPException(status_code=400, detail=f"Unknown data source: {source_id}")

    dashboard_dir = app_config.dashboards_dir() / slug

    result = await execute_dashboard(
        config=config,
        filter_values=request.filters,
        pagination={},
        data_source_filter=source_id,
        dashboard_dir=dashboard_dir,
    )

    ds = result.data_sources.get(source_id, {})
    if isinstance(ds, dict) and ds.get("status") == "error":
        raise HTTPException(status_code=502, detail=ds.get("error", "Query failed"))

    rows = ds.get("rows", []) if isinstance(ds, dict) else []
    if len(rows) > 100_000:
        raise HTTPException(status_code=400, detail="Export limited to 100,000 rows. Use filters to narrow results.")

    columns = ds.get("columns", []) if isinstance(ds, dict) else []

    # Get column definitions for formatting
    col_defs = {}
    for node in config.layout:
        if node.data_source == source_id and node.columns:
            for col in node.columns:
                col_defs[col.key] = col
            break

    wb = Workbook()
    ws = wb.active
    ws.title = source_id

    # Header row (bold)
    bold = Font(bold=True)
    for col_idx, col_key in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_defs[col_key].label if col_key in col_defs else col_key
        cell.font = bold

    # Format map
    fmt_map = {
        "currency": '#,##0.00',
        "number": '#,##0',
        "date": 'YYYY-MM-DD',
        "datetime": 'YYYY-MM-DD HH:MM:SS',
    }

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, col_key in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = row.get(col_key, "")
            col_def = col_defs.get(col_key)
            col_fmt = col_def.format if col_def else None

            if col_fmt in ("number", "currency") and val not in (None, "", "--"):
                try:
                    cell.value = float(val)
                    if col_fmt in fmt_map:
                        cell.number_format = fmt_map[col_fmt]
                except (ValueError, TypeError):
                    cell.value = str(val)
            else:
                cell.value = str(val) if val is not None else ""

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    rate_limiter.record(user_id)
    filename = f"{slug}_{source_id}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/dashboard/{slug}/config")
async def dashboard_config(
    slug: str,
    claims: dict = Depends(require_auth),
):
    """Return dashboard config without SQL content."""
    perms = get_user_permissions(claims)
    if not can_access_dashboard(perms, slug):
        raise HTTPException(status_code=403, detail="Dashboard access denied")

    try:
        config = load_dashboard_config(slug)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"Dashboard '{slug}' not found")

    # Return config as dict, sql_ref fields are kept (frontend ignores them)
    return config.model_dump()


@app.post("/api/dashboard/{slug}/page")
async def dashboard_page(
    slug: str,
    request: PageRequest,
    claims: dict = Depends(require_auth),
) -> PageResponseItem:
    """Paginate a single data source within a dashboard."""
    user_id = claims.get("sub", "unknown")
    allowed, reason = rate_limiter.check(user_id)
    if not allowed:
        raise HTTPException(status_code=429, detail=reason)

    perms = get_user_permissions(claims)
    if not can_access_dashboard(perms, slug):
        raise HTTPException(status_code=403, detail="Dashboard access denied")

    try:
        config = load_dashboard_config(slug)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"Dashboard '{slug}' not found")

    # Validate data source exists
    if request.data_source not in config.data_sources:
        raise HTTPException(status_code=400, detail=f"Unknown data source: {request.data_source}")

    source = config.data_sources[request.data_source]

    # Depth cap: max offset 50000
    offset = (max(request.page, 1) - 1) * request.page_size
    if offset > 50000:
        raise HTTPException(status_code=400, detail="Page too deep -- use filters to narrow results")

    # Validate sort column if provided
    if request.sort_column:
        from .dashboard import get_panel_columns
        allowed_cols = get_panel_columns(config, request.data_source)
        if allowed_cols and request.sort_column not in allowed_cols:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid sort column '{request.sort_column}'. Allowed: {sorted(allowed_cols)}",
            )

    from .dashboard_query import build_filtered_query
    from .duckdb_engine import get_duckdb_engine
    from .dashboard_models import DuckDBSource, SortConfig
    from datetime import datetime, timezone

    if not isinstance(source, DuckDBSource):
        raise HTTPException(status_code=400, detail="Pagination only supported for DuckDB sources")

    dashboard_dir = app_config.dashboards_dir() / slug
    sql_path = dashboard_dir / source.sql_ref
    if not sql_path.exists():
        raise HTTPException(status_code=404, detail=f"SQL file not found: {source.sql_ref}")

    base_sql = sql_path.read_text()

    sort = SortConfig(column=request.sort_column, direction=request.sort_direction) if request.sort_column else None
    from .dashboard import get_panel_columns
    allowed_cols = get_panel_columns(config, request.data_source)

    all_bindings = source.filters + source.text_filters
    filtered_sql, params = build_filtered_query(
        base_sql=base_sql,
        filter_bindings=all_bindings,
        filter_values=request.filters,
        filter_defs=config.filters,
        sort=sort,
        page=request.page,
        page_size=request.page_size,
        allowed_sort_columns=allowed_cols if allowed_cols else None,
    )

    engine = get_duckdb_engine()

    # Count query
    count_sql = f"WITH base AS ({base_sql}) SELECT COUNT(*) as total FROM base"
    try:
        count_result = engine.query(count_sql, {})
        total_rows = count_result["rows"][0]["total"] if count_result["rows"] else 0
    except Exception:
        total_rows = 0

    # Execute paginated query
    try:
        result = engine.query(filtered_sql, params)
        columns = result["columns"]
        result_rows = result["rows"]
    except Exception as e:
        rate_limiter.record(user_id)
        return PageResponseItem(
            status="error",
            error=str(e),
            page=request.page,
            page_size=request.page_size,
            queried_at=datetime.now(timezone.utc).isoformat(),
        )

    rate_limiter.record(user_id)
    return PageResponseItem(
        status="ok",
        columns=columns,
        rows=result_rows,
        total_rows=total_rows,
        page=request.page,
        page_size=request.page_size,
        queried_at=datetime.now(timezone.utc).isoformat(),
    )


# ---------------------------------------------------------------------------
# Validate dashboard configs at startup
# ---------------------------------------------------------------------------

class RequestIDMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        rid = str(uuid.uuid4())
        app_config.request_id.set(rid)
        structlog.contextvars.bind_contextvars(request_id=rid)

        user_id = ""
        if hasattr(request.state, "user_id"):
            user_id = request.state.user_id

        log = structlog.get_logger("request")
        log.info("request.start", method=request.method, path=request.url.path, user_id=user_id)

        start = time.perf_counter()
        response = await call_next(request)
        duration_ms = round((time.perf_counter() - start) * 1000, 1)

        log.info(
            "request.complete",
            method=request.method,
            path=request.url.path,
            status=response.status_code,
            duration_ms=duration_ms,
            user_id=user_id,
        )
        response.headers["X-Request-ID"] = rid
        structlog.contextvars.unbind_contextvars("request_id")
        return response


app.add_middleware(RequestIDMiddleware)


class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    CSP = "; ".join([
        "default-src 'self'",
        "script-src 'self' https://clerk.com https://*.clerk.accounts.dev https://*.clerk.dev",
        "style-src 'self' 'unsafe-inline'",
        "connect-src 'self' https://clerk.com https://*.clerk.accounts.dev https://*.clerk.dev",
        "img-src 'self' data: https://*.clerk.com",
        "font-src 'self'",
        "frame-ancestors 'none'",
    ])

    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        response.headers["Content-Security-Policy-Report-Only"] = self.CSP
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
        response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
        return response


app.add_middleware(SecurityHeadersMiddleware)


@app.on_event("startup")
async def validate_dashboards_on_startup():
    """Validate config, init logging, init DuckDB, validate dashboards."""
    app_config.setup_logging()
    app_config.validate()

    from .duckdb_engine import init_duckdb_engine

    try:
        engine = init_duckdb_engine()
        logger.info(
            "Startup complete",
            extra={
                "environment": app_config.ENVIRONMENT,
                "content_root": str(app_config.CONTENT_ROOT),
                "tables_loaded": list(engine._tables.keys()),
            },
        )
    except Exception as e:
        logger.warning(f"DuckDB engine init failed (dashboard queries will fail): {e}")

    results = validate_all_dashboards()
    for slug, errors in results.items():
        if errors:
            for err in errors:
                logger.warning(f"Dashboard config warning [{slug}]: {err}")


@app.post("/api/admin/reload-config")
async def admin_reload_config(claims: dict = Depends(require_auth)):
    """Validate all dashboard configs and confirm they are loadable.

    Admin-only. Validates every config before reporting success — on any
    validation error the response reflects those errors and no swap occurs.
    If all configs are clean, returns the list of loaded dashboards.
    """
    perms = get_user_permissions(claims)
    if not is_admin(perms):
        raise HTTPException(status_code=403, detail="Admin access required")

    user_id = claims.get("sub", "unknown")
    user_email = claims.get("email", "")

    validation_results = validate_all_dashboards()

    errors_by_slug = {slug: errs for slug, errs in validation_results.items() if errs}
    loaded = [slug for slug in validation_results if not validation_results[slug]]

    event_logger.log(
        "admin.reload_config",
        {
            "user_id": user_id,
            "user_email": user_email,
            "dashboards_checked": list(validation_results.keys()),
            "dashboards_loaded": loaded,
            "validation_errors": errors_by_slug,
            "success": len(errors_by_slug) == 0,
        },
    )

    if errors_by_slug:
        return {
            "success": False,
            "dashboards_loaded": [],
            "validation_errors": errors_by_slug,
        }

    return {
        "success": True,
        "dashboards_loaded": loaded,
        "validation_errors": {},
    }


@app.get("/api/dashboards")
async def list_dashboards(claims: dict = Depends(require_auth)):
    """List dashboards the current user can access."""
    perms = get_user_permissions(claims)
    registry = get_dashboard_registry()

    if "*" in perms.allowed_dashboards:
        visible_slugs = list(registry.keys())
    else:
        visible_slugs = [s for s in perms.allowed_dashboards if s in registry]

    dashboards = [
        {
            "slug": slug,
            "title": registry[slug].get("title", slug),
            "description": registry[slug].get("description", ""),
        }
        for slug in visible_slugs
    ]
    return {"dashboards": dashboards}


# ---------------------------------------------------------------------------
# Dashboard static file serving (with cookie auth)
# ---------------------------------------------------------------------------

@app.get("/api/permissions")
async def get_permissions(claims: dict = Depends(require_auth)):
    """Return the authenticated user's dashboard permissions."""
    perms = get_user_permissions(claims)
    registry = get_dashboard_registry()

    if "*" in perms.allowed_dashboards:
        visible_slugs = list(registry.keys())
    else:
        visible_slugs = [s for s in perms.allowed_dashboards if s in registry]

    dashboards = [
        {
            "slug": slug,
            "title": registry[slug]["title"],
            "description": registry[slug]["description"],
            "path": registry[slug]["path"],
        }
        for slug in visible_slugs
    ]

    return {"groups": perms.groups, "dashboards": dashboards}


# Mount content data directory (before SPA catch-all)
_content_data_mount = app_config.CONTENT_ROOT / "data"
if _content_data_mount.exists():
    app.mount("/data", StaticFiles(directory=str(_content_data_mount)))

# Serve static files in production (SPA catch-all — must be last)
static_dir = Path(__file__).parent.parent / "static"
if static_dir.exists():
    from starlette.responses import FileResponse

    # Mount /assets for hashed static files (JS, CSS)
    assets_dir = static_dir / "assets"
    if assets_dir.exists():
        app.mount("/assets", StaticFiles(directory=str(assets_dir)))

    # SPA catch-all: serve index.html for all non-API routes
    @app.get("/{full_path:path}")
    async def spa_catch_all(full_path: str):
        # Serve actual static files if they exist (favicon, vite.svg, etc.)
        file_path = static_dir / full_path
        if full_path and file_path.exists() and file_path.is_file():
            return FileResponse(str(file_path))
        return FileResponse(str(static_dir / "index.html"))
